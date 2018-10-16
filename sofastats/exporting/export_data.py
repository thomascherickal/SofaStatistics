from abc import ABC, abstractmethod
from collections import namedtuple
import csv
import datetime
from pathlib import Path

import openpyxl #@UnresolvedImport
from openpyxl.styles import Alignment, Font
import wx

from sofastats import basic_lib as b
from sofastats import my_globals as mg
from sofastats import lib
from sofastats import my_exceptions
from sofastats import getdata
from sofastats import output

ColDet = namedtuple('ColDet', 'col_name, numeric, is_date, valdic')


class Writer(ABC):

    """
    Write out content of table to output e.g. CSV. Optionally include label
    fields at end.
    """

    def __init__(self, cols_dets, *, inc_lbls):
        """
        :param list cols_dets: list of column details to use
        :param bool inc_lbls: if True, include label versions of fields at end
         as extra columns.
        """
        self.dd = mg.DATADETS_OBJ
        self.col_names = []
        for col_dets in cols_dets:
            col_name = col_dets.col_name
            if col_name == mg.SOFA_ID:
                col_name = mg.WAS_SOFA_ID
            self.col_names.append(col_name)
            if col_dets.valdic:
                self.col_names.append(f'{col_name}_Label')
        self.extra_lbl = '_with_labels' if inc_lbls else ''
        self.desktop = Path(mg.HOME_PATH) / 'Desktop'

    @abstractmethod
    def process_row(self, row, row_n, cols_dets):
        """
        Iterate through items in row. If item has a label variant to display
        then also add that item to the output.
        """
        pass

    @abstractmethod
    def save(self):
        pass


class CsvWriter(Writer):

    def __init__(self, cols_dets, *, inc_lbls):
        super().__init__(cols_dets, inc_lbls=inc_lbls)
        fname_csv = f'{self.dd.tbl}{self.extra_lbl}.csv'
        fpath_csv = str(self.desktop / fname_csv)
        self.f = open(fpath_csv, 'w')
        self.csv_writer = csv.writer(
            self.f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        self.csv_writer.writerow(self.col_names)  ## write header row

    def process_row(self, row, row_n, cols_dets):
        try:
            self.csv_writer.writerow(row)
        except Exception as e:
            raise Exception(f'Unable to write row {row_n} to csv file. '
                f'Orig error: {e}')

    def save(self):
        self.f.close()


class SpreadsheetWriter(Writer):

    @staticmethod
    def _get_date_col_nums(cols_dets):
        date_col_nums = []
        n = 1
        for col_dets in cols_dets:
            if col_dets.is_date:
                date_col_nums.append(n)
            if col_dets.valdic:
                n += 1  ## need to increment by one because there will be a lbl col before the next orig col (if any)
            n += 1  ## the next orig col
        return date_col_nums

    def __init__(self, cols_dets, *, inc_lbls):
        super().__init__(cols_dets, inc_lbls=inc_lbls)
        self.wb = openpyxl.Workbook(write_only=False)  ## write_only faster but can't add rows cell-by-cell, only by appending entire row
        safe_sheetname = lib.get_safer_name(f'{self.dd.tbl[:20]} output')
        self.sheet = self.wb.create_sheet(safe_sheetname, 0)
        date_col_nums = SpreadsheetWriter._get_date_col_nums(cols_dets)
        style_bold_12pt = Font(name='Arial', size=12, bold=True)
        for col_num, col_name in enumerate(self.col_names, 1):
            header_cell = self.sheet.cell(row=1, column=col_num)
            header_cell.value = col_name
            header_cell.font = style_bold_12pt
        ## Have to wait till columns are filled with something before accessing sheet.columns reliably
        for col_num, _unused in enumerate(self.col_names, 1):
            if col_num in date_col_nums:
                ## https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
                cols = list(self.sheet.columns)
                col_idx = col_num - 1
                col = cols[col_idx]
                column_name = col[0].column
                self.sheet.column_dimensions[column_name].width = 13

    def process_row(self, row, row_n, cols_dets):
        try:
            col2fill = 1
            for val, col_dets in zip(row, cols_dets):
                cell = self.sheet.cell(row=row_n, column=col2fill)
                if col_dets.is_date:
                    if isinstance(val, datetime.datetime):
                        val2use = val
                    else:
                        try:
                            val2use = lib.DateLib.get_datetime_from_str(val)
                        except Exception:
                            val2use = ''
                    cell.value = val2use
                    cell.number_format = 'YYYY MMM DD'
                else:
                    if col_dets.numeric:
                        cell.alignment = Alignment(horizontal='right')
                    cell.value = val
                if col_dets.valdic:
                    col2fill += 1
                    val2show = col_dets.valdic.get(val, val)
                    self.sheet.cell(row=row_n, column=col2fill).value = val2show
                col2fill += 1
        except Exception as e:
            raise Exception(f'Unable to write row {row_n} to spreadsheet. '
                f'Orig error: {e}')

    def save(self):
        fname_xlsx = f'{self.dd.tbl}{self.extra_lbl}.xlsx'
        fpath_xlsx = str(self.desktop / fname_xlsx)
        self.wb.save(fpath_xlsx)


class DlgExportData(wx.Dialog):

    def __init__(self, n_rows):
        """
        temp_desktop_report_only -- exporting output to a temporary desktop
        folder

        Works with SQLite, PostgreSQL, MySQL, CUBRID, MS Access and SQL Server.
        """
        self.dd = mg.DATADETS_OBJ
        title = f'Export {self.dd.tbl} to spreadsheet'
        wx.Dialog.__init__(self, parent=None, id=-1, title=title,
            pos=(mg.HORIZ_OFFSET+200, 300),
            style=(wx.MINIMIZE_BOX|wx.MAXIMIZE_BOX|wx.RESIZE_BORDER|wx.CLOSE_BOX
                |wx.SYSTEM_MENU|wx.CAPTION|wx.CLIP_CHILDREN))
        szr = wx.BoxSizer(wx.VERTICAL)
        self.n_rows = n_rows
        self.export_status = {mg.CANCEL_EXPORT: False}  ## can change and running script can check on it.
        self.chk_inc_lbls = wx.CheckBox(
            self, -1, _('Include extra label columns (if available)'))
        szr.Add(self.chk_inc_lbls, 0, wx.ALL, 10)
        self.btn_cancel = wx.Button(self, wx.ID_CANCEL)
        self.btn_cancel.Bind(wx.EVT_BUTTON, self.on_btn_cancel)
        self.btn_cancel.Enable(False)
        self.btn_export = wx.Button(self, -1, 'Export')
        self.btn_export.Bind(wx.EVT_BUTTON, self.on_btn_export)
        self.btn_export.Enable(True)
        szr_btns = wx.FlexGridSizer(rows=1, cols=2, hgap=5, vgap=0)
        szr_btns.AddGrowableCol(1,2)  ## idx, propn
        szr_btns.Add(self.btn_cancel, 0)
        szr_btns.Add(self.btn_export, 0, wx.ALIGN_RIGHT)
        self.progbar = wx.Gauge(self, -1, mg.EXPORT_DATA_GAUGE_STEPS,
            size=(-1, 20), style=wx.GA_HORIZONTAL)
        self.btn_close = wx.Button(self, wx.ID_CLOSE)
        self.btn_close.Bind(wx.EVT_BUTTON, self.on_btn_close)
        szr.Add(szr_btns, 0, wx.GROW|wx.LEFT|wx.RIGHT|wx.BOTTOM, 10)
        szr.Add(self.progbar, 0, wx.GROW|wx.ALIGN_RIGHT|wx.ALL, 10)
        szr.Add(self.btn_close, 0, wx.ALIGN_RIGHT|wx.ALL, 10)
        self.SetSizer(szr)
        szr.SetSizeHints(self)
        szr.Layout()

    def get_cols_dets(self, orig_col_names, *, val_dics, inc_lbls):
        """
        Get details for each column e.g. whether numeric, and whether it has a
        label to be displayed as well as then original value

        :return: list of ColDet named tuples: col_name, numeric, is_date, valdic
        :rtype: list
        """
        debug = False
        cols_dets = []
        for orig_colname in orig_col_names:
            numericfld = self.dd.flds[orig_colname][mg.FLD_BOLNUMERIC]
            datefld = self.dd.flds[orig_colname][mg.FLD_BOLDATETIME]
            if debug and datefld: print(f'{orig_colname} is a datetime field')
            has_lbls = val_dics.get(orig_colname)
            use_lbls = (inc_lbls and has_lbls)
            lbl_dic = val_dics.get(orig_colname) if use_lbls else None
            cols_dets.append(ColDet(orig_colname, numericfld, datefld, lbl_dic))
        return cols_dets

    def _get_sql(self, orig_col_names):
        debug = False
        objqtr = getdata.get_obj_quoter_func(self.dd.dbe)
        colnames_clause = ', '.join([objqtr(x) for x in orig_col_names])
        dbe_tblname = getdata.tblname_qtr(self.dd.dbe, self.dd.tbl)
        sql_get_data = f'SELECT {colnames_clause} FROM {dbe_tblname}'
        if debug: print(sql_get_data)
        return sql_get_data

    def write_out(self, cols_dets, *, inc_lbls, do_spreadsheet):
        """
        For each writer (CSV and, optionally, spreadsheet) write out content.
        """
        writers = [CsvWriter(cols_dets, inc_lbls=inc_lbls), ]
        if do_spreadsheet:
            writers.append(SpreadsheetWriter(cols_dets, inc_lbls=inc_lbls))
        orig_col_names = [col_dets.col_name for col_dets in cols_dets]
        sql_get_data = self._get_sql(orig_col_names)
        self.dd.cur.execute(sql_get_data)  ## must be dd.cur
        row_n = 2  ## one-based and already added header row
        while True:
            row = self.dd.cur.fetchone()  ## Note - won't have the extra columns, just the original columns
            if row is None:
                break
            for writer in writers:
                writer.process_row(row, row_n, cols_dets)
            gauge2set = min(
                (row_n/self.n_rows)*mg.EXPORT_DATA_GAUGE_STEPS,
                mg.EXPORT_DATA_GAUGE_STEPS)
            self.progbar.SetValue(gauge2set)
            row_n += 1
        for writer in writers:
            writer.save()

    def on_btn_export(self, _evt):
        """
        Make CSV and, if possible, a spreadsheet as well (must not have too many
        columns).

        Follows format of original data source where possible.

        Start with original list of columns and column names. If including
        labels, with insert additional fields as required.
        """
        debug = False
        self.progbar.SetValue(0)
        wx.BeginBusyCursor()
        inc_lbls = self.chk_inc_lbls.IsChecked()
        orig_col_names = getdata.fldsdic_to_fldnames_lst(fldsdic=self.dd.flds)
        cc = output.get_cc()
        (_var_labels, _var_notes,
         _var_types, val_dics) = lib.get_var_dets(cc[mg.CURRENT_VDTS_PATH])
        cols_dets = self.get_cols_dets(
            orig_col_names, val_dics=val_dics, inc_lbls=inc_lbls)
        n_used_lbls = len(
            [col_dets for col_dets in cols_dets if col_dets.valdic])
        n_cols = len(cols_dets) + n_used_lbls
        do_spreadsheet = (n_cols <= 256)
        if not do_spreadsheet:
            wx.MessageBox(f'Too many columns ({n_cols}) for an '
                'xlsx spreadsheet. Only making the csv')
        try:
            self.write_out(
                cols_dets, inc_lbls=inc_lbls, do_spreadsheet=do_spreadsheet)
        except my_exceptions.ExportCancel:
            wx.MessageBox('Export Cancelled')
        except Exception as e:
            msg = (f'Problem exporting output. Orig error: {b.ue(e)}')
            if debug: print(msg)
            wx.MessageBox(msg)
            self.progbar.SetValue(0)
            lib.GuiLib.safe_end_cursor()
            self.align_btns_to_exporting(exporting=False)
            self.export_status[mg.CANCEL_EXPORT] = False
            return
        self.progbar.SetValue(mg.EXPORT_DATA_GAUGE_STEPS)
        lib.GuiLib.safe_end_cursor()
        self.align_btns_to_exporting(exporting=False)
        wx.MessageBox(_('Your data has been exported to your desktop'))
        self.progbar.SetValue(0)

    def on_btn_cancel(self, _evt):
        self.export_status[mg.CANCEL_EXPORT] = True

    def align_btns_to_exporting(self, exporting):
        self.btn_close.Enable(not exporting)
        self.btn_cancel.Enable(exporting)
        self.btn_export.Enable(not exporting)

    def on_btn_close(self, _evt):
        self.Destroy()
