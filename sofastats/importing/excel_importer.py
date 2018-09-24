import wx
import openpyxl

from sofastats import basic_lib as b  #@UnresolvedImport
from sofastats import my_globals as mg  #@UnresolvedImport
from sofastats import my_exceptions  #@UnresolvedImport
from sofastats import lib  #@UnresolvedImport
from sofastats import getdata  #@UnresolvedImport
from sofastats.importing import importer  #@UnresolvedImport

ROWS_TO_SAMPLE = 500  ## fast enough to sample quite a few


class ExcelImporter(importer.FileImporter):
    """
    Import excel file into default SOFA SQLite database.

    Needs to identify data types to ensure only consistent data in a field.

    Adds unique index id so can identify unique records with certainty.
    """

    def __init__(self, parent, fpath, tblname, *,
            headless, headless_has_header, force_quickcheck=False):
        importer.FileImporter.__init__(self, parent, fpath, tblname,
            headless=headless, headless_has_header=headless_has_header)
        self.ext = 'XLS/XLSX'
        self.force_quickcheck = force_quickcheck

    def _get_start_row_idx(self):
        start_row_idx = 2 if self.has_header else 1  ## openpyxl is 1-based
        return start_row_idx

    def has_header_row(self, row1_types, row2_types):
        """
        Will return True if nothing but strings or in first row and anything in
        other rows that is not e.g. a number or a date. Empty is OK.
        """
        debug = False
        if debug: print(row1_types, row2_types)
        str_type = openpyxl.cell.cell.Cell.TYPE_STRING  #@UndefinedVariable
        empty_type = openpyxl.cell.cell.Cell.TYPE_NULL  #@UndefinedVariable
        non_str_types = [
            openpyxl.cell.cell.Cell.TYPE_BOOL,  #@UndefinedVariable
            openpyxl.cell.cell.Cell.TYPE_NUMERIC,  #@UndefinedVariable
        ]  ## ignore TYPE_FORMULA, TYPE_ERROR, TYPE_FORMULA_CACHE_STRING, TYPE_INLINE
        return importer.has_header_row(
            row1_types, row2_types, str_type, empty_type, non_str_types)

    @staticmethod
    def _get_processed_val(raw_val, none_replacement):
        """
        Get processed version of raw value ready for HTML renderer to display or
        SQLite to ingest. Strings OK for both cases.

        :param unknown raw_val: raw value. Need to remove trailing space from
         strings and handle nulls (None).
        :param str none_replacement: OK to be a string because SQLite will
         handle it appropriately according to its affinity e.g. if affinity is
         numeric then "3" will be stored as numeric; or HTML will display it
         depending on use case.
        :return: processed version of raw val
        :rtype: str
        """
        if raw_val is None:
            val = none_replacement
        else:
            try:
                val = raw_val.strip()
            except AttributeError:
                val = str(raw_val)
        return val

    def get_params(self):
        """
        Display each cell based on cell characteristics only - no attempt to
        collapse down to one data type e.g. dates only or numbers only.
        """
        debug = False
        if self.headless:
            self.has_header = self.headless_has_header
            return True
        else:
            wkbook = openpyxl.load_workbook(self.fpath)
            wksheet = wkbook.worksheets[0]
            n_cols = wksheet.max_column
            n_rows = wksheet.max_row
            strdata = []
            row1_types = []
            row2_types = []
            for row_idx in range(1, n_rows + 1):
                new_row = []
                for col_idx in range(1, n_cols + 1):
                    cell = wksheet.cell(row=row_idx, column=col_idx)
                    raw_val = cell.value
                    data_type = cell.data_type
                    if row_idx == 1:
                        row1_types.append(data_type)
                    elif row_idx == 2:
                        row2_types.append(data_type)
                    val = ExcelImporter._get_processed_val(
                        raw_val, none_replacement='&nbsp;'*3)  ## wide enough to see cell if all cells in sample column are empty
                    new_row.append(val)
                strdata.append(new_row)
                if (row_idx + 1) >= importer.ROWS_TO_SHOW_USER:
                    break
            try:
                prob_has_hdr = self.has_header_row(row1_types, row2_types)
            except Exception:
                prob_has_hdr = False
            dlg = importer.DlgHasHeaderGivenData(
                self.parent, self.ext, strdata, prob_has_hdr)
            ret = dlg.ShowModal()
            if debug: print(str(ret))
            if ret == wx.ID_CANCEL:
                return False
            else:
                self.has_header = (ret == mg.HAS_HEADER)
            return True

    def get_ok_fldnames(self, wksheet):
        n_cols = wksheet.max_column
        if self.has_header:
            ## use values of first row
            orig_fldnames = []
            for col_idx in range(1, n_cols + 1):
                raw_fldname = wksheet.cell(row=1, column=col_idx).value
                fldname = (raw_fldname if isinstance(raw_fldname, str) 
                    else str(raw_fldname))
                orig_fldnames.append(fldname)
            fldnames = importer.process_fldnames(orig_fldnames, 
                headless=self.headless, force_quickcheck=self.force_quickcheck)
        else:
            ## numbered is OK
            fldnames = [
                mg.NEXT_FLDNAME_TEMPLATE % (x + 1, ) for x in range(n_cols)]
        return fldnames

    def get_rowdict(self, row_idx, wksheet, fldnames):
        row_vals = []
        n_cols = wksheet.max_column
        for col_idx in range(1, n_cols + 1):
            raw_val = wksheet.cell(row=row_idx, column=col_idx).value
            val = ExcelImporter._get_processed_val(raw_val, none_replacement='')
            row_vals.append(val)
        rowdict = dict(zip(fldnames, row_vals))
        return rowdict

    def assess_sample(self, wksheet, ok_fldnames,
            progbar, steps_per_item, import_status,
            faulty2missing_fld_list):
        """
        Assess data sample to identify field types based on values in fields.
        Sample first N rows (at most) to establish field types. If a field has
        mixed data types will define as string.

        https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

        :param dict fldtypes: dict with ok field names as keys and field types
         as values.
        :param list sample_data: list of dicts containing the first rows of data
         (no point reading them all again during subsequent steps).
        :return: fldtypes, sample_data.
        :rtype: tuple
        """
        debug = False
        has_rows = False
        sample_data = []
        n_rows = wksheet.max_row
        row_idx = self._get_start_row_idx()
        while row_idx <= n_rows:  ## iterates through data rows only
            if row_idx % 50 == 0:
                if not self.headless:
                    wx.Yield()
                if import_status[mg.CANCEL_IMPORT]:
                    progbar.SetValue(0)
                    raise my_exceptions.ImportCancel
            if debug: print(self.get_rowdict(row_idx, wksheet, ok_fldnames))
            ## if has_header, starts at 1st data row
            has_rows = True
            rowdict = self.get_rowdict(row_idx, wksheet, ok_fldnames)
            sample_data.append(rowdict)
            gauge_val = min(row_idx*steps_per_item, mg.IMPORT_GAUGE_STEPS)
            progbar.SetValue(gauge_val)
            if not self.headless:
                if row_idx == ROWS_TO_SAMPLE:
                    break
            row_idx += 1
        fldtypes = []
        for ok_fldname in ok_fldnames:
            fldtype = importer.assess_sample_fld(sample_data,
                ok_fldname, ok_fldnames, faulty2missing_fld_list,
                has_header=self.has_header, headless=self.headless)
            fldtypes.append(fldtype)            
        fldtypes = dict(zip(ok_fldnames, fldtypes))
        if not has_rows:
            raise Exception('No data to import')
        return fldtypes, sample_data

    def import_content(self,
            lbl_feedback=None, import_status=None, progbar=None):
        """
        Get field types dict.  Use it to test each and every item before they
        are added to database (after adding the records already tested).

        Add to disposable table first and if completely successful, rename
        table to final name.
        """
        debug = False
        if lbl_feedback is None: lbl_feedback = importer.DummyLabel()
        if import_status is None:
            import_status = importer.dummy_import_status.copy()
        if progbar is None: progbar = importer.DummyProgBar()
        faulty2missing_fld_list = []
        if not self.headless:
            wx.BeginBusyCursor()
        try:
            wkbook = openpyxl.load_workbook(self.fpath)
            wksheet = wkbook.worksheets[0]
            n_rows = wksheet.max_row
            n_datarows = n_rows - 1 if self.has_header else n_rows
            ## get field names
            ok_fldnames = self.get_ok_fldnames(wksheet)
            if debug: print(ok_fldnames)
        except OSError as e:
            lib.GuiLib.safe_end_cursor()
            raise Exception(
                f'Unable to find file "{self.fpath}" for importing.')
        except Exception as e:
            lib.GuiLib.safe_end_cursor()
            raise Exception(
                f'Unable to read spreadsheet.\nCaused by error: {b.ue(e)}')
        default_dd = getdata.get_default_db_dets()
        if self.headless:
            global ROWS_TO_SAMPLE
            ROWS_TO_SAMPLE = n_datarows
        sample_n = min(ROWS_TO_SAMPLE, n_datarows)
        items_n = n_datarows + sample_n
        steps_per_item = importer.get_steps_per_item(items_n)
        if debug: 
            print(f'steps_per_item: {steps_per_item}')
            print('About to assess data sample')
        fldtypes, sample_data = self.assess_sample(wksheet, ok_fldnames,
            progbar, steps_per_item, import_status,
            faulty2missing_fld_list)
        if debug:
            print('Just finished assessing data sample')
            print(fldtypes)
            print(sample_data)
        data = []
        row_idx = self._get_start_row_idx()
        while row_idx <= n_rows:  ## iterates through data rows only
            data.append(
                self.get_rowdict(row_idx, wksheet, ok_fldnames))
            row_idx += 1
        gauge_start = steps_per_item * sample_n
        try:
            feedback = {mg.NULLED_DOTS_KEY: False}
            importer.add_to_tmp_tbl(
                feedback, import_status,
                default_dd.con, default_dd.cur,
                self.tblname, ok_fldnames, fldtypes,
                faulty2missing_fld_list, data,
                progbar, n_datarows, steps_per_item, gauge_start,
                has_header=self.has_header, headless=self.headless)
            importer.tmp_to_named_tbl(default_dd.con, default_dd.cur,
                self.tblname, progbar, feedback[mg.NULLED_DOTS_KEY],
                headless=self.headless)
        except Exception as e:
            importer.post_fail_tidy(progbar, default_dd.con, default_dd.cur)
            raise
        default_dd.cur.close()
        default_dd.con.commit()
        default_dd.con.close()
        progbar.SetValue(0)
